import streamlit as st
import pandas as pd
import numpy as np
import re
import openpyxl
import xlrd
import sqlite3
import hashlib
import uuid
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from io import BytesIO, StringIO
import plotly.graph_objects as go
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors

# ====================== 基础配置 ======================
st.set_page_config(page_title="集装箱配箱系统｜终极版", page_icon="📦", layout="wide")
st.title("📦 集装箱智能配箱系统｜3D可视化+PDF版")

# ====================== 数据库初始化 ======================
def init_db():
    conn = sqlite3.connect("container_system.db")
    c = conn.cursor()
    # 用户表
    c.execute('''CREATE TABLE IF NOT EXISTS users 
                 (id TEXT PRIMARY KEY, username TEXT UNIQUE, password TEXT, role TEXT)''')
    # 模板表
    c.execute('''CREATE TABLE IF NOT EXISTS templates 
                 (id TEXT PRIMARY KEY, name TEXT, type TEXT, config TEXT, create_time TEXT)''')
    # 批次表
    c.execute('''CREATE TABLE IF NOT EXISTS batches 
                 (id TEXT PRIMARY KEY, user_id TEXT, name TEXT, data TEXT, create_time TEXT)''')
    # 日志表
    c.execute('''CREATE TABLE IF NOT EXISTS logs 
                 (id TEXT PRIMARY KEY, user_id TEXT, action TEXT, create_time TEXT)''')
    # 默认管理员
    c.execute("SELECT * FROM users WHERE username='admin'")
    if not c.fetchone():
        c.execute("INSERT INTO users VALUES (?,?,?,?)", 
                  (str(uuid.uuid4()), "admin", hashlib.md5("admin123".encode()).hexdigest(), "admin"))
    conn.commit()
    conn.close()

init_db()

# ====================== 权限控制 ======================
if "user" not in st.session_state:
    st.session_state.user = None

def login(username, password):
    conn = sqlite3.connect("container_system.db")
    c = conn.cursor()
    c.execute("SELECT * FROM users WHERE username=? AND password=?", 
              (username, hashlib.md5(password.encode()).hexdigest()))
    user = c.fetchone()
    conn.close()
    return user

def log_action(user_id, action):
    conn = sqlite3.connect("container_system.db")
    c = conn.cursor()
    c.execute("INSERT INTO logs VALUES (?,?,?,?)", 
              (str(uuid.uuid4()), user_id, action, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    conn.commit()
    conn.close()

# ====================== 箱型定义 ======================
CONTAINER_SPECS = {
    "20GP": {"long": 5898, "width": 2352, "height": 2393, "max_weight": 28000, "tare": 2200},
    "40GP": {"long": 12032, "width": 2352, "height": 2393, "max_weight": 30480, "tare": 3750},
    "40HQ": {"long": 12032, "width": 2352, "height": 2698, "max_weight": 30480, "tare": 4000},
}

# ====================== 全局状态 ======================
keys = ["cargo_df", "alloc_result", "pack_3d_data", "template_import", "template_export"]
defaults = [
    pd.DataFrame(columns=["货物名称","长(mm)","宽(mm)","高(mm)","毛重(kg)","净重(kg)","柜号","来源"]),
    {}, {}, None, None
]
for k, v in zip(keys, defaults):
    if k not in st.session_state:
        st.session_state[k] = v

# ====================== 1. Excel解析引擎（xls+xlsx+微软+WPS） ======================
class UltimateExcelParser:
    def __init__(self, file, filename):
        self.file = file
        self.filename = filename
        self.items = []

    def clean(self, v):
        return str(v).strip().replace("\n","") if v is not None else ""

    def parse_val(self, val, ctx=""):
        s = self.clean(val).lower() + self.clean(ctx).lower()
        nums = re.findall(r"[0-9.]+", s)
        v = float(nums[0]) if nums else 0.0
        if "cm" in s: return v*10
        if "m" in s: return v*1000
        if "g" in s: return v/1000
        if "t" in s or "吨" in s: return v*1000
        return v

    def parse_xlsx(self):
        try:
            wb = load_workbook(self.file, data_only=True)
            for sn in wb.sheetnames:
                ws = wb[sn]
                for r in range(1, min(ws.max_row+1, 2000)):
                    name = self.clean(ws.cell(r,1).value)
                    if not name: continue
                    self.items.append({
                        "货物名称": name,
                        "长(mm)": round(self.parse_val(ws.cell(r,2).value),2),
                        "宽(mm)": round(self.parse_val(ws.cell(r,3).value),2),
                        "高(mm)": round(self.parse_val(ws.cell(r,4).value),2),
                        "毛重(kg)": round(self.parse_val(ws.cell(r,5).value),2),
                        "净重(kg)": round(self.parse_val(ws.cell(r,6).value),2),
                        "柜号": "", "来源": f"{sn}行{r}"
                    })
            return True
        except:
            return False

    def parse_xls(self):
        try:
            wb = xlrd.open_workbook(file_contents=self.file.read())
            for s in wb.sheets():
                for r in range(min(s.nrows, 2000)):
                    row = s.row_values(r)
                    if len(row)<3: continue
                    name = self.clean(row[0])
                    if not name: continue
                    self.items.append({
                        "货物名称": name,
                        "长(mm)": round(self.parse_val(row[1] if len(row)>1 else 0),2),
                        "宽(mm)": round(self.parse_val(row[2] if len(row)>2 else 0),2),
                        "高(mm)": round(self.parse_val(row[3] if len(row)>3 else 0),2),
                        "毛重(kg)": round(self.parse_val(row[4] if len(row)>4 else 0),2),
                        "净重(kg)": round(self.parse_val(row[5] if len(row)>5 else 0),2),
                        "柜号": "", "来源": s.name
                    })
            return True
        except:
            return False

    def parse(self):
        if self.filename.endswith(".xlsx"): self.parse_xlsx()
        elif self.filename.endswith(".xls"): self.parse_xls()
        return self.items

# ====================== 2. 3D装箱算法 + 可视化 ======================
class Pack3D:
    def __init__(self, container_type):
        self.spec = CONTAINER_SPECS[container_type]
        self.available = [(0, 0, 0, self.spec["long"], self.spec["width"], self.spec["height"])]
        self.loaded_weight = 0
        self.packed = []

    def pack_item(self, item):
        l, w, h = item["长(mm)"], item["宽(mm)"], item["高(mm)"]
        weight = item["毛重(kg)"]
        
        if self.loaded_weight + weight > (self.spec["max_weight"] - self.spec["tare"]):
            return False

        for i, space in enumerate(self.available):
            sx, sy, sz, sl, sw, sh = space
            # 6种旋转方式
            for rot in [(l,w,h), (l,h,w), (w,l,h), (w,h,l), (h,l,w), (h,w,l)]:
                rl, rw, rh = rot
                if rl <= sl and rw <= sw and rh <= sh:
                    self.packed.append({
                        "name": item["货物名称"],
                        "pos": (sx, sy, sz),
                        "size": (rl, rw, rh),
                        "weight": weight
                    })
                    self.loaded_weight += weight

                    # 更新可用空间
                    new_spaces = []
                    if sl - rl > 10:
                        new_spaces.append((sx+rl, sy, sz, sl-rl, sw, sh))
                    if sw - rw > 10:
                        new_spaces.append((sx, sy+rw, sz, rl, sw-rw, sh))
                    if sh - rh > 10:
                        new_spaces.append((sx, sy, sz+rh, rl, rw, sh-rh))
                    
                    del self.available[i]
                    self.available.extend(new_spaces)
                    return True
        return False

def multi_box_pack(df, types):
    res = {}
    pack_3d = {}
    df = df.copy()
    df["柜号"] = ""
    
    for container_type in types:
        packer = Pack3D(container_type)
        box_no = 1
        for i, row in df.iterrows():
            if df.at[i, "柜号"]: continue
            item = row.to_dict()
            if not packer.pack_item(item):
                box_no +=1
                packer = Pack3D(container_type)
                if not packer.pack_item(item):
                    continue  # 单个货物超重/超体积
            code = f"{container_type}{box_no}"
            df.at[i, "柜号"] = code
            if code not in res:
                res[code] = {"箱型":container_type, "总重":0, "总件":0}
            res[code]["总重"] += row["毛重(kg)"]
            res[code]["总件"] +=1
            pack_3d[code] = packer.packed
    
    st.session_state.pack_3d_data = pack_3d
    return df, res

# 3D可视化渲染
def render_3d_packing(container_code):
    if container_code not in st.session_state.pack_3d_data:
        return go.Figure()
    
    packed = st.session_state.pack_3d_data[container_code]
    container_type = container_code[:4] if container_code.startswith("40HQ") else container_code[:3]
    spec = CONTAINER_SPECS[container_type]
    
    fig = go.Figure()
    
    # 绘制集装箱箱体
    fig.add_trace(go.Mesh3d(
        x=[0, spec["long"], spec["long"], 0, 0, spec["long"], spec["long"], 0],
        y=[0, 0, spec["width"], spec["width"], 0, 0, spec["width"], spec["width"]],
        z=[0, 0, 0, 0, spec["height"], spec["height"], spec["height"], spec["height"]],
        i=[0, 1, 2, 0, 4, 5, 6, 4],
        j=[1, 2, 3, 3, 5, 6, 7, 7],
        k=[2, 3, 0, 1, 6, 7, 4, 5],
        opacity=0.1,
        color="lightblue",
        name="集装箱"
    ))
    
    # 绘制货物
    colors_list = ["red", "green", "blue", "yellow", "purple", "orange", "pink", "brown"]
    for i, item in enumerate(packed):
        x0, y0, z0 = item["pos"]
        l, w, h = item["size"]
        color = colors_list[i % len(colors_list)]
        
        fig.add_trace(go.Mesh3d(
            x=[x0, x0+l, x0+l, x0, x0, x0+l, x0+l, x0],
            y=[y0, y0, y0+w, y0+w, y0, y0, y0+w, y0+w],
            z=[z0, z0, z0, z0, z0+h, z0+h, z0+h, z0+h],
            i=[0, 1, 2, 0, 4, 5, 6, 4],
            j=[1, 2, 3, 3, 5, 6, 7, 7],
            k=[2, 3, 0, 1, 6, 7, 4, 5],
            opacity=0.6,
            color=color,
            name=item["name"][:8] + "..." if len(item["name"]) > 8 else item["name"]
        ))
    
    fig.update_layout(
        scene=dict(
            xaxis_title="长度 (mm)",
            yaxis_title="宽度 (mm)",
            zaxis_title="高度 (mm)",
            aspectmode="data"
        ),
        title=f"{container_code} 3D装箱可视化",
        width=800,
        height=600
    )
    return fig

# ====================== 3. PDF装箱单生成 ======================
def generate_packing_pdf(df, container_code=None):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    style_title = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=20,
        alignment=1
    )
    
    elements = []
    elements.append(Paragraph("集装箱装箱单", style_title))
    elements.append(Paragraph(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]))
    elements.append(Paragraph("="*50, styles["Normal"]))
    
    # 筛选指定柜号数据
    if container_code:
        df_filtered = df[df["柜号"] == container_code]
        elements.append(Paragraph(f"柜号：{container_code}", styles["Heading2"]))
    else:
        df_filtered = df
    
    # 生成表格
    data = [["货物名称", "长(mm)", "宽(mm)", "高(mm)", "毛重(kg)", "净重(kg)", "柜号"]]
    for _, row in df_filtered.iterrows():
        data.append([
            row["货物名称"],
            row["长(mm)"],
            row["宽(mm)"],
            row["高(mm)"],
            row["毛重(kg)"],
            row["净重(kg)"],
            row["柜号"]
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.grey),
        ("TEXTCOLOR", (0,0), (-1,0), colors.whitesmoke),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0,0), (-1,0), 12),
        ("GRID", (0,0), (-1,-1), 1, colors.black),
    ]))
    elements.append(table)
    
    # 汇总信息
    elements.append(Paragraph("\n汇总信息：", styles["Heading2"]))
    summary = df_filtered.groupby("柜号").agg({
        "货物名称": "count",
        "毛重(kg)": "sum"
    }).rename(columns={"货物名称": "总件数", "毛重(kg)": "总重量(kg)"})
    
    summary_data = [["柜号", "总件数", "总重量(kg)"]]
    for idx, row in summary.iterrows():
        summary_data.append([idx, row["总件数"], row["总重量(kg)"]])
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.lightblue),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("GRID", (0,0), (-1,-1), 1, colors.black),
    ]))
    elements.append(summary_table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer

# ====================== 登录界面 ======================
if not st.session_state.user:
    with st.form("login_form"):
        st.subheader("🔐 系统登录")
        username = st.text_input("用户名")
        password = st.text_input("密码", type="password")
        if st.form_submit_button("登录"):
            user = login(username, password)
            if user:
                st.session_state.user = {
                    "id": user[0],
                    "name": user[1],
                    "role": user[3]
                }
                log_action(user[0], "用户登录")
                st.success("登录成功！")
                st.rerun()
            else:
                st.error("用户名或密码错误！")
    st.stop()

# ====================== 主界面 ======================
st.sidebar.title(f"欢迎 {st.session_state.user['name']}")
menu = st.sidebar.radio("功能菜单", [
    "货物导入", "3D配箱计算", "装箱单/PDF", "3D可视化", "批次管理", "操作日志"
])

# 退出登录
if st.sidebar.button("🚪 退出登录"):
    log_action(st.session_state.user["id"], "用户退出")
    st.session_state.user = None
    st.rerun()

# ---------------------- 1. 货物导入 ----------------------
if menu == "货物导入":
    st.subheader("📥 Excel导入（支持.xls/.xlsx，微软/WPS兼容）")
    uploaded_file = st.file_uploader("上传货物清单", type=["xls", "xlsx"])
    
    if uploaded_file and st.button("✅ 解析导入"):
        with st.spinner("双引擎解析中..."):
            parser = UltimateExcelParser(uploaded_file, uploaded_file.name)
            data = parser.parse()
            st.session_state.cargo_df = pd.DataFrame(data)
            log_action(st.session_state.user["id"], f"导入{len(data)}条货物数据")
        st.success(f"导入成功！共解析 {len(data)} 条有效货物")
    
    st.dataframe(st.session_state.cargo_df, use_container_width=True)

# ---------------------- 2. 3D配箱计算 ----------------------
elif menu == "3D配箱计算":
    st.subheader("🚀 3D多箱型混合配箱")
    if st.session_state.cargo_df.empty:
        st.warning("请先导入货物数据！")
    else:
        selected_types = st.multiselect(
            "选择可用箱型", 
            list(CONTAINER_SPECS.keys()), 
            default=["20GP", "40HQ"]
        )
        if st.button("开始3D配箱计算"):
            with st.spinner("3D装箱计算中..."):
                df_out, alloc_result = multi_box_pack(st.session_state.cargo_df, selected_types)
                st.session_state.cargo_df = df_out
                st.session_state.alloc_result = alloc_result
                log_action(st.session_state.user["id"], "执行3D配箱计算")
            st.success("配箱计算完成！")
        
        # 显示配箱结果
        st.dataframe(st.session_state.cargo_df, use_container_width=True)
        if st.session_state.alloc_result:
            st.subheader("配箱汇总")
            summary_df = pd.DataFrame([
                {"柜号":k, **v} for k, v in st.session_state.alloc_result.items()
            ])
            st.dataframe(summary_df, use_container_width=True)

# ---------------------- 3. 装箱单/PDF ----------------------
elif menu == "装箱单/PDF":
    st.subheader("📄 装箱单生成 & PDF导出")
    if st.session_state.cargo_df.empty:
        st.warning("请先完成配箱计算！")
    else:
        container_list = st.session_state.cargo_df["柜号"].unique()
        selected_container = st.selectbox(
            "选择柜号（留空导出全部）", 
            [""] + list(container_list)
        )
        
        # 生成PDF
        if st.button("生成PDF装箱单"):
            with st.spinner("PDF生成中..."):
                pdf_buffer = generate_packing_pdf(
                    st.session_state.cargo_df, 
                    selected_container if selected_container else None
                )
                log_action(st.session_state.user["id"], f"生成{selected_container or '全部'}PDF装箱单")
            
            filename = f"装箱单_{selected_container or '全部'}_{datetime.now().strftime('%Y%m%d')}.pdf"
            st.download_button(
                label="📥 下载PDF装箱单",
                data=pdf_buffer,
                file_name=filename,
                mime="application/pdf"
            )
            
            # 预览
            st.subheader("装箱单预览")
            preview_df = st.session_state.cargo_df[
                st.session_state.cargo_df["柜号"] == selected_container
            ] if selected_container else st.session_state.cargo_df
            st.dataframe(preview_df, use_container_width=True)

# ---------------------- 4. 3D可视化 ----------------------
elif menu == "3D可视化":
    st.subheader("🎯 3D装箱可视化")
    if not st.session_state.pack_3d_data:
        st.warning("请先完成3D配箱计算！")
    else:
        container_code = st.selectbox(
            "选择柜号查看3D装箱",
            list(st.session_state.pack_3d_data.keys())
        )
        if container_code:
            fig = render_3d_packing(container_code)
            st.plotly_chart(fig, use_container_width=True)

# ---------------------- 5. 批次管理 ----------------------
elif menu == "批次管理":
    st.subheader("📊 配箱批次管理")
    batch_name = st.text_input("输入批次名称")
    if batch_name and st.button("💾 保存当前批次"):
        conn = sqlite3.connect("container_system.db")
        c = conn.cursor()
        c.execute("INSERT INTO batches VALUES (?,?,?,?,?)", 
                  (str(uuid.uuid4()), 
                   st.session_state.user["id"],
                   batch_name,
                   st.session_state.cargo_df.to_json(),
                   datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        conn.commit()
        conn.close()
        log_action(st.session_state.user["id"], f"保存批次：{batch_name}")
        st.success(f"批次「{batch_name}」保存成功！")
    
    # 显示历史批次
    conn = sqlite3.connect("container_system.db")
    batch_df = pd.read_sql(
        "SELECT * FROM batches WHERE user_id=? ORDER BY create_time DESC",
        conn,
        params=(st.session_state.user["id"],)
    )
    conn.close()
    st.dataframe(batch_df[["id", "name", "create_time"]], use_container_width=True)

# ---------------------- 6. 操作日志 ----------------------
elif menu == "操作日志":
    st.subheader("📜 操作日志")
    if st.session_state.user["role"] == "admin":
        conn = sqlite3.connect("container_system.db")
        log_df = pd.read_sql("SELECT * FROM logs ORDER BY create_time DESC", conn)
        conn.close()
        st.dataframe(log_df, use_container_width=True)
    else:
        st.warning("⚠️ 仅管理员可查看操作日志！")

st.caption("✅ 终极版：3D可视化｜PDF装箱单｜Excel全兼容｜数据库存储｜权限控制")
